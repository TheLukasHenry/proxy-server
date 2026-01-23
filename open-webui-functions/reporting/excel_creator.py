"""
title: Excel Creator
author: MCP Team
author_url: https://github.com/your-org
funding_url: https://github.com/sponsors/your-org
version: 0.1.0
description: Creates Excel files with working formulas, multiple sheets, and professional formatting. Outputs downloadable .xlsx files.
requirements: openpyxl
"""

import json
import base64
import io
import re
from datetime import datetime
from typing import Optional, Any


class Tools:
    """
    Excel Creator Tool for Open WebUI

    Creates professional Excel spreadsheets with:
    - Working formulas (=SUM, =AVERAGE, =IF, =VLOOKUP, etc.)
    - Multiple sheets
    - Formatting (bold, currency, percentages, dates)
    - Auto-fit columns
    - Freeze panes

    Example prompt:
        "Create an Excel file with Q1 sales data:
         - Products: Widget, Gadget, Gizmo
         - January sales: 100, 150, 200
         - February sales: 120, 140, 180
         - Add a Total column with SUM formula
         - Add a totals row at the bottom"

    The LLM should call this tool with a JSON specification.
    """

    def __init__(self):
        self.valves = self.Valves()

    class Valves:
        """Configuration options for the Excel Creator tool."""
        pass

    def create_excel(
        self,
        specification: str,
        __user__: dict = {},
        __event_emitter__: Any = None
    ) -> str:
        """
        Create an Excel file with formulas and formatting.

        :param specification: JSON specification for the Excel file. Format:
            {
                "filename": "report_name",
                "sheets": [
                    {
                        "name": "Sheet Name",
                        "headers": ["Col1", "Col2", "Col3", "Total"],
                        "data": [
                            ["Row1", 100, 200, "=SUM(B2:C2)"],
                            ["Row2", 150, 250, "=SUM(B3:C3)"]
                        ],
                        "totals_row": ["TOTAL", "=SUM(B2:B3)", "=SUM(C2:C3)", "=SUM(D2:D3)"],
                        "format": {
                            "bold_headers": true,
                            "currency_columns": ["B", "C", "D"],
                            "percentage_columns": [],
                            "freeze_row": 1,
                            "auto_fit": true
                        }
                    }
                ]
            }
        :return: HTML with download trigger or error message
        """
        try:
            # Import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                from openpyxl.utils import get_column_letter
            except ImportError:
                return "Error: openpyxl library not installed. Please install it with: pip install openpyxl"

            # Parse specification
            if isinstance(specification, str):
                try:
                    spec = json.loads(specification)
                except json.JSONDecodeError as e:
                    return f"Error: Invalid JSON specification. {str(e)}"
            else:
                spec = specification

            # Validate required fields
            if "sheets" not in spec or not spec["sheets"]:
                return "Error: Specification must include at least one sheet in 'sheets' array."

            # Create workbook
            wb = Workbook()
            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)

            # Define styles
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Process each sheet
            for sheet_spec in spec["sheets"]:
                sheet_name = sheet_spec.get("name", "Sheet")[:31]  # Excel limit
                ws = wb.create_sheet(title=sheet_name)

                format_spec = sheet_spec.get("format", {})
                row_num = 1

                # Write headers
                headers = sheet_spec.get("headers", [])
                if headers:
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=header)
                        if format_spec.get("bold_headers", True):
                            cell.font = header_font
                            cell.fill = header_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
                    row_num += 1

                # Write data rows
                data = sheet_spec.get("data", [])
                for row_data in data:
                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.border = thin_border

                        # Apply number formatting
                        col_letter = get_column_letter(col_num)
                        if col_letter in format_spec.get("currency_columns", []):
                            cell.number_format = '$#,##0.00'
                        elif col_letter in format_spec.get("percentage_columns", []):
                            cell.number_format = '0.00%'
                    row_num += 1

                # Write totals row
                totals_row = sheet_spec.get("totals_row", [])
                if totals_row:
                    for col_num, value in enumerate(totals_row, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.font = Font(bold=True)
                        cell.border = thin_border

                        # Apply number formatting to totals
                        col_letter = get_column_letter(col_num)
                        if col_letter in format_spec.get("currency_columns", []):
                            cell.number_format = '$#,##0.00'
                        elif col_letter in format_spec.get("percentage_columns", []):
                            cell.number_format = '0.00%'

                # Auto-fit columns
                if format_spec.get("auto_fit", True):
                    for column in ws.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                cell_value = str(cell.value) if cell.value else ""
                                # For formulas, estimate display length
                                if cell_value.startswith("="):
                                    cell_value = "000000.00"  # Estimate for numbers
                                if len(cell_value) > max_length:
                                    max_length = len(cell_value)
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50
                        ws.column_dimensions[column_letter].width = adjusted_width

                # Freeze panes
                freeze_row = format_spec.get("freeze_row", 0)
                if freeze_row > 0:
                    ws.freeze_panes = f"A{freeze_row + 1}"

            # Save to bytes buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            excel_bytes = buffer.read()

            # Encode as base64
            b64_content = base64.b64encode(excel_bytes).decode('utf-8')

            # Generate filename
            base_filename = spec.get("filename", "excel_report")
            # Sanitize filename
            safe_filename = re.sub(r'[^\w\-]', '_', base_filename)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            full_filename = f"{safe_filename}_{timestamp}.xlsx"

            # Count sheets and rows for summary
            total_sheets = len(spec["sheets"])
            total_rows = sum(len(s.get("data", [])) for s in spec["sheets"])

            # Return HTML that triggers download
            return f'''
<div style="padding: 20px; background: #f0f9ff; border-radius: 8px; border: 1px solid #0ea5e9;">
    <h3 style="margin: 0 0 10px 0; color: #0369a1;">Excel File Created Successfully</h3>
    <p style="margin: 0 0 15px 0; color: #475569;">
        <strong>File:</strong> {full_filename}<br>
        <strong>Sheets:</strong> {total_sheets}<br>
        <strong>Data Rows:</strong> {total_rows}
    </p>
    <p style="margin: 0; color: #64748b; font-size: 12px;">
        The file includes working formulas that will calculate automatically in Excel.
    </p>
</div>

<script>
(function() {{
    try {{
        const b64 = "{b64_content}";
        const byteCharacters = atob(b64);
        const byteNumbers = new Array(byteCharacters.length);
        for (let i = 0; i < byteCharacters.length; i++) {{
            byteNumbers[i] = byteCharacters.charCodeAt(i);
        }}
        const byteArray = new Uint8Array(byteNumbers);
        const blob = new Blob([byteArray], {{
            type: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }});
        const url = URL.createObjectURL(blob);
        const a = document.createElement("a");
        a.href = url;
        a.download = "{full_filename}";
        document.body.appendChild(a);
        a.click();
        document.body.removeChild(a);
        URL.revokeObjectURL(url);
    }} catch (e) {{
        console.error("Download failed:", e);
    }}
}})();
</script>
'''

        except Exception as e:
            return f"Error creating Excel file: {str(e)}"

    def create_simple_excel(
        self,
        title: str,
        headers: str,
        data: str,
        include_totals: bool = True,
        __user__: dict = {},
        __event_emitter__: Any = None
    ) -> str:
        """
        Create a simple Excel file from basic inputs (easier for LLM to use).

        :param title: Title/filename for the Excel file
        :param headers: Comma-separated column headers (e.g., "Product,Jan,Feb,Mar,Total")
        :param data: Pipe-separated rows, comma-separated values (e.g., "Widget,100,120,140|Gadget,150,160,170")
        :param include_totals: Whether to add a totals row with SUM formulas
        :return: HTML with download trigger
        """
        try:
            # Parse headers
            header_list = [h.strip() for h in headers.split(",")]

            # Parse data rows
            rows = []
            for row_str in data.split("|"):
                row_values = []
                for val in row_str.split(","):
                    val = val.strip()
                    # Try to convert to number
                    try:
                        if "." in val:
                            row_values.append(float(val))
                        else:
                            row_values.append(int(val))
                    except ValueError:
                        row_values.append(val)
                rows.append(row_values)

            # Detect which columns are numeric (for totals)
            numeric_cols = []
            if rows:
                for col_idx in range(len(rows[0])):
                    if all(isinstance(row[col_idx], (int, float)) for row in rows if col_idx < len(row)):
                        numeric_cols.append(col_idx)

            # Build specification
            spec = {
                "filename": title,
                "sheets": [{
                    "name": title[:31],
                    "headers": header_list,
                    "data": rows,
                    "format": {
                        "bold_headers": True,
                        "auto_fit": True,
                        "freeze_row": 1
                    }
                }]
            }

            # Add totals row if requested
            if include_totals and numeric_cols:
                totals_row = []
                data_row_count = len(rows)
                start_row = 2  # After header
                end_row = start_row + data_row_count - 1

                for col_idx in range(len(header_list)):
                    if col_idx == 0:
                        totals_row.append("TOTAL")
                    elif col_idx in numeric_cols:
                        col_letter = chr(65 + col_idx)  # A=65
                        totals_row.append(f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
                    else:
                        totals_row.append("")

                spec["sheets"][0]["totals_row"] = totals_row

            # Call the main function
            return self.create_excel(json.dumps(spec), __user__, __event_emitter__)

        except Exception as e:
            return f"Error creating simple Excel: {str(e)}"
