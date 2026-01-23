#!/usr/bin/env python3
"""Insert Excel Creator as a proper Tool (not Function) in Open WebUI."""
import json
import time
import os

DB_URL = os.environ.get("DATABASE_URL", "postgresql://openwebui:localdev@postgres:5432/openwebui")

tool_id = "excel_creator"
user_id = "8a2851d8-3aa9-4963-a987-a71df3bc40db"

# Tool content with class Tools
content = '''import base64
import io


class Tools:
    def __init__(self):
        pass

    def create_excel(self, title: str, headers: str, data: str) -> str:
        """
        Create an Excel file with data and SUM formulas. Returns an HTML download link.

        :param title: Name for the Excel file (e.g., "Q1Sales")
        :param headers: Comma-separated column headers (e.g., "Product,Jan,Feb,Mar")
        :param data: Pipe-separated rows with comma-separated values (e.g., "Widget,100,150,200|Gadget,250,300,350")
        :return: HTML with styled download link for the Excel file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
        except ImportError:
            return "<p style='color:red;'>Error: openpyxl not installed</p>"

        # Parse headers
        header_list = [h.strip() for h in headers.split(",")]

        # Parse data rows
        rows = []
        for row_str in data.split("|"):
            row = []
            for val in row_str.split(","):
                val = val.strip()
                try:
                    val = int(val)
                except ValueError:
                    try:
                        val = float(val)
                    except ValueError:
                        pass
                row.append(val)
            rows.append(row)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = str(title)[:31]

        # Write headers
        for col, header in enumerate(header_list, 1):
            ws.cell(row=1, column=col, value=str(header)).font = Font(bold=True)

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Add totals row
        total_row = len(rows) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        for col in range(2, len(header_list) + 1):
            col_letter = get_column_letter(col)
            formula = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
            ws.cell(row=total_row, column=col, value=formula).font = Font(bold=True)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Encode as base64
        b64 = base64.b64encode(buffer.read()).decode()
        fname = f"{title}.xlsx"

        # Return HTML with download link
        return f"""<div style="padding:20px;background:linear-gradient(135deg,#1e3a5f,#2d5a87);border-radius:12px;color:white;margin:10px 0;">
<h3 style="margin:0 0 10px 0;">📊 Excel File Created: {fname}</h3>
<p style="margin:0 0 15px 0;opacity:0.9;">Contains {len(rows)} data rows with {len(header_list)} columns and SUM formulas</p>
<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{fname}" style="display:inline-block;padding:12px 24px;background:#4CAF50;color:white;text-decoration:none;border-radius:6px;font-weight:bold;">⬇️ Download {fname}</a>
</div>"""
'''

# OpenAPI spec for the tool
specs = [
    {
        "name": "create_excel",
        "description": "Create an Excel file with data and SUM formulas. Returns an HTML download link.",
        "parameters": {
            "type": "object",
            "properties": {
                "title": {
                    "type": "string",
                    "description": "Name for the Excel file (e.g., 'Q1Sales')"
                },
                "headers": {
                    "type": "string",
                    "description": "Comma-separated column headers (e.g., 'Product,Jan,Feb,Mar')"
                },
                "data": {
                    "type": "string",
                    "description": "Pipe-separated rows with comma-separated values (e.g., 'Widget,100,150,200|Gadget,250,300,350')"
                }
            },
            "required": ["title", "headers", "data"]
        }
    }
]

meta = {"description": "Creates Excel files with data and SUM formulas"}
now = int(time.time())

print(f"Connecting to PostgreSQL...")

try:
    import psycopg2
except ImportError:
    import subprocess
    subprocess.run(["pip", "install", "psycopg2-binary", "-q"])
    import psycopg2

conn = psycopg2.connect(DB_URL)
cursor = conn.cursor()

# Delete from function table if exists
cursor.execute('DELETE FROM function WHERE id = %s', (tool_id,))
print(f"Cleaned up function table")

# Delete existing tool if exists
cursor.execute('DELETE FROM tool WHERE id = %s', (tool_id,))

# Insert as tool
print(f"Inserting tool {tool_id}...")
cursor.execute('''
    INSERT INTO tool (id, user_id, name, content, specs, meta, created_at, updated_at, valves, access_control)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
''', (
    tool_id,
    user_id,
    "Excel Creator",
    content,
    json.dumps(specs),
    json.dumps(meta),
    now,
    now,
    '{}',
    json.dumps({"read": {"group_ids": [], "user_ids": []}, "write": {"group_ids": [], "user_ids": []}})
))

conn.commit()
print(f"Inserted! Rows affected: {cursor.rowcount}")

# Verify
cursor.execute('SELECT id, name FROM tool')
rows = cursor.fetchall()
print(f"Tools in database: {rows}")

conn.close()
print("Done!")
