#!/usr/bin/env python3
"""Insert Excel Creator function into PostgreSQL (runs inside Docker)."""
import json
import time
import os

# Get DATABASE_URL from environment
DB_URL = os.environ.get("DATABASE_URL", "postgresql://openwebui:localdev@postgres:5432/openwebui")

function_id = "excel_creator"
user_id = "8a2851d8-3aa9-4963-a987-a71df3bc40db"
name = "Excel Creator"
func_type = "tool"

content = '''import base64
import io


class Tools:
    def __init__(self):
        pass

    def create_excel(self, title: str, headers: str, data: str) -> str:
        """
        Create an Excel file with data and SUM formulas.

        :param title: Name for the Excel file
        :param headers: Comma-separated headers like Product,Jan,Feb,Mar
        :param data: Pipe-separated rows like Widget,100,120,90|Gadget,150,140,160
        :return: HTML with download link
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
        except ImportError:
            return "Error: openpyxl not installed"

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]

        header_list = [h.strip() for h in headers.split(",")]
        for col, header in enumerate(header_list, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        row_num = 2
        for row_str in data.split("|"):
            col_num = 1
            for val in row_str.split(","):
                val = val.strip()
                try:
                    val = int(val)
                except ValueError:
                    pass
                ws.cell(row=row_num, column=col_num, value=val)
                col_num += 1
            row_num += 1

        ws.cell(row=row_num, column=1, value="TOTAL").font = Font(bold=True)
        for col in range(2, len(header_list) + 1):
            col_letter = get_column_letter(col)
            formula = "=SUM(" + col_letter + "2:" + col_letter + str(row_num - 1) + ")"
            ws.cell(row=row_num, column=col, value=formula).font = Font(bold=True)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        b64 = base64.b64encode(buffer.read()).decode()
        fname = title + ".xlsx"

        return "<a href='data:application/octet-stream;base64," + b64 + "' download='" + fname + "'>Download " + fname + "</a>"
'''

meta = json.dumps({"description": "Creates Excel files with data and formulas"})
now = int(time.time())

print(f"Connecting to PostgreSQL at {DB_URL}...")

try:
    import psycopg2
except ImportError:
    print("Installing psycopg2...")
    import subprocess
    subprocess.run(["pip", "install", "psycopg2-binary", "-q"])
    import psycopg2

conn = psycopg2.connect(DB_URL)
cursor = conn.cursor()

# Delete if exists
cursor.execute('DELETE FROM function WHERE id = %s', (function_id,))
print(f"Deleted existing function (if any)")

# Insert
print(f"Inserting function {function_id}...")
cursor.execute('''
    INSERT INTO function (id, user_id, name, type, content, meta, created_at, updated_at, valves, is_active, is_global)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
''', (function_id, user_id, name, func_type, content, meta, now, now, '{}', True, True))

conn.commit()
print(f"Committed!")

# Verify
cursor.execute('SELECT id, name, type, is_active FROM function')
rows = cursor.fetchall()
print(f"Functions in database: {rows}")

conn.close()
print("Done!")
