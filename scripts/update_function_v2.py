#!/usr/bin/env python3
"""Update Excel Creator function to handle all input formats."""
import json
import time
import os

DB_URL = os.environ.get("DATABASE_URL", "postgresql://openwebui:localdev@postgres:5432/openwebui")

function_id = "excel_creator"

# Updated content that handles ALL formats the model might send
content = '''import base64
import io


class Tools:
    def __init__(self):
        pass

    def create_excel(self, title: str, headers, data) -> str:
        """
        Create an Excel file with data and SUM formulas.

        :param title: Name for the Excel file (e.g., "Q1Sales")
        :param headers: Column headers as list ["Product","Jan","Feb","Mar"] or comma-separated string
        :param data: Row data as nested list [["Widget",100,150,200]] or list of comma-separated strings ["Widget,100,150,200"]
        :return: HTML with download link for the Excel file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
        except ImportError:
            return "<p style='color:red;'>Error: openpyxl not installed</p>"

        # Handle headers - can be string or list
        if isinstance(headers, str):
            header_list = [h.strip() for h in headers.split(",")]
        elif isinstance(headers, list):
            header_list = [str(h) for h in headers]
        else:
            return f"<p style='color:red;'>Error: headers must be string or list, got {type(headers)}</p>"

        # Handle data - can be string, list of strings, or nested list
        rows = []
        if isinstance(data, str):
            # Pipe-separated format: "Widget,100,150,200|Gadget,250,300,350"
            for row_str in data.split("|"):
                row = []
                for val in row_str.split(","):
                    val = val.strip()
                    try:
                        val = int(val)
                    except ValueError:
                        try:
                            val = float(val)
                        except ValueError:
                            pass
                    row.append(val)
                rows.append(row)
        elif isinstance(data, list):
            for item in data:
                if isinstance(item, str):
                    # List of comma-separated strings: ["Widget,100,150,200", "Gadget,250,300,350"]
                    row = []
                    for val in item.split(","):
                        val = val.strip()
                        try:
                            val = int(val)
                        except ValueError:
                            try:
                                val = float(val)
                            except ValueError:
                                pass
                        row.append(val)
                    rows.append(row)
                elif isinstance(item, list):
                    # Nested list: [["Widget",100,150,200], ["Gadget",250,300,350]]
                    rows.append(item)
                else:
                    rows.append([item])
        else:
            return f"<p style='color:red;'>Error: data must be string or list, got {type(data)}</p>"

        if not rows:
            return "<p style='color:red;'>Error: no data rows provided</p>"

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = str(title)[:31]

        # Write headers
        for col, header in enumerate(header_list, 1):
            ws.cell(row=1, column=col, value=str(header)).font = Font(bold=True)

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Add totals row with SUM formulas for numeric columns
        total_row = len(rows) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)

        for col in range(2, len(header_list) + 1):
            col_letter = get_column_letter(col)
            formula = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
            ws.cell(row=total_row, column=col, value=formula).font = Font(bold=True)

        # Save to bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Encode as base64
        b64 = base64.b64encode(buffer.read()).decode()
        fname = f"{title}.xlsx"

        # Return HTML with styled box and download link
        return f"""<div style="padding:20px;background:linear-gradient(135deg,#1e3a5f,#2d5a87);border-radius:12px;color:white;margin:10px 0;">
<h3 style="margin:0 0 10px 0;">📊 Excel File Created: {fname}</h3>
<p style="margin:0 0 15px 0;opacity:0.9;">Contains {len(rows)} data rows with {len(header_list)} columns</p>
<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{fname}" style="display:inline-block;padding:12px 24px;background:#4CAF50;color:white;text-decoration:none;border-radius:6px;font-weight:bold;">⬇️ Download {fname}</a>
</div>"""
'''

print(f"Connecting to PostgreSQL...")

try:
    import psycopg2
except ImportError:
    print("Installing psycopg2...")
    import subprocess
    subprocess.run(["pip", "install", "psycopg2-binary", "-q"])
    import psycopg2

conn = psycopg2.connect(DB_URL)
cursor = conn.cursor()

now = int(time.time())

# Update the function
print(f"Updating function {function_id}...")
cursor.execute('''
    UPDATE function SET content = %s, updated_at = %s WHERE id = %s
''', (content, now, function_id))

conn.commit()
print(f"Updated! Rows affected: {cursor.rowcount}")

conn.close()
print("Done!")
