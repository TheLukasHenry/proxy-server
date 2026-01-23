#!/usr/bin/env python3
"""Insert full Reporting Toolkit (Excel Creator + Executive Dashboard) into Open WebUI."""
import json
import time
import os

DB_URL = os.environ.get("DATABASE_URL", "postgresql://openwebui:localdev@postgres:5432/openwebui")

user_id = "8a2851d8-3aa9-4963-a987-a71df3bc40db"
now = int(time.time())

# ============================================================================
# TOOL 1: Excel Creator (Full Version)
# ============================================================================

excel_creator_id = "excel_creator"
excel_creator_name = "Excel Creator"

excel_creator_content = '''import json
import base64
import io
import re
from datetime import datetime
from typing import Optional, Any


class Tools:
    def __init__(self):
        pass

    def create_excel(
        self,
        specification: str,
        __user__: dict = {},
        __event_emitter__: Any = None
    ) -> str:
        """
        Create an Excel file with formulas and formatting.

        :param specification: JSON specification for the Excel file. Format:
            {
                "filename": "report_name",
                "sheets": [
                    {
                        "name": "Sheet Name",
                        "headers": ["Col1", "Col2", "Col3", "Total"],
                        "data": [
                            ["Row1", 100, 200, "=SUM(B2:C2)"],
                            ["Row2", 150, 250, "=SUM(B3:C3)"]
                        ],
                        "totals_row": ["TOTAL", "=SUM(B2:B3)", "=SUM(C2:C3)", "=SUM(D2:D3)"],
                        "format": {
                            "bold_headers": true,
                            "currency_columns": ["B", "C", "D"],
                            "percentage_columns": [],
                            "freeze_row": 1,
                            "auto_fit": true
                        }
                    }
                ]
            }
        :return: HTML with download trigger or error message
        """
        try:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                from openpyxl.utils import get_column_letter
            except ImportError:
                return "<p style='color:red;'>Error: openpyxl not installed</p>"

            if isinstance(specification, str):
                try:
                    spec = json.loads(specification)
                except json.JSONDecodeError as e:
                    return f"<p style='color:red;'>Error: Invalid JSON - {str(e)}</p>"
            else:
                spec = specification

            if "sheets" not in spec or not spec["sheets"]:
                return "<p style='color:red;'>Error: Must include at least one sheet</p>"

            wb = Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)

            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for sheet_spec in spec["sheets"]:
                sheet_name = sheet_spec.get("name", "Sheet")[:31]
                ws = wb.create_sheet(title=sheet_name)

                format_spec = sheet_spec.get("format", {})
                row_num = 1

                headers = sheet_spec.get("headers", [])
                if headers:
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=header)
                        if format_spec.get("bold_headers", True):
                            cell.font = header_font
                            cell.fill = header_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
                    row_num += 1

                data = sheet_spec.get("data", [])
                for row_data in data:
                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.border = thin_border
                        col_letter = get_column_letter(col_num)
                        if col_letter in format_spec.get("currency_columns", []):
                            cell.number_format = '$#,##0.00'
                        elif col_letter in format_spec.get("percentage_columns", []):
                            cell.number_format = '0.00%'
                    row_num += 1

                totals_row = sheet_spec.get("totals_row", [])
                if totals_row:
                    for col_num, value in enumerate(totals_row, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.font = Font(bold=True)
                        cell.border = thin_border
                        col_letter = get_column_letter(col_num)
                        if col_letter in format_spec.get("currency_columns", []):
                            cell.number_format = '$#,##0.00'
                        elif col_letter in format_spec.get("percentage_columns", []):
                            cell.number_format = '0.00%'

                if format_spec.get("auto_fit", True):
                    for column in ws.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                cell_value = str(cell.value) if cell.value else ""
                                if cell_value.startswith("="):
                                    cell_value = "000000.00"
                                if len(cell_value) > max_length:
                                    max_length = len(cell_value)
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width

                freeze_row = format_spec.get("freeze_row", 0)
                if freeze_row > 0:
                    ws.freeze_panes = f"A{freeze_row + 1}"

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            excel_bytes = buffer.read()

            b64_content = base64.b64encode(excel_bytes).decode('utf-8')

            base_filename = spec.get("filename", "excel_report")
            safe_filename = re.sub(r'[^\\w\\-]', '_', base_filename)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            full_filename = f"{safe_filename}_{timestamp}.xlsx"

            total_sheets = len(spec["sheets"])
            total_rows = sum(len(s.get("data", [])) for s in spec["sheets"])

            return f"""<div style="padding:20px;background:linear-gradient(135deg,#1e3a5f,#2d5a87);border-radius:12px;color:white;margin:10px 0;">
<h3 style="margin:0 0 10px 0;">Excel File Created: {full_filename}</h3>
<p style="margin:0 0 10px 0;opacity:0.9;">Sheets: {total_sheets} | Data Rows: {total_rows} | Formulas: Active</p>
<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_content}" download="{full_filename}" style="display:inline-block;padding:12px 24px;background:#4CAF50;color:white;text-decoration:none;border-radius:6px;font-weight:bold;">Download {full_filename}</a>
</div>"""

        except Exception as e:
            return f"<p style='color:red;'>Error: {str(e)}</p>"

    def create_simple_excel(
        self,
        title: str,
        headers: str,
        data: str,
        include_totals: bool = True,
        __user__: dict = {},
        __event_emitter__: Any = None
    ) -> str:
        """
        Create a simple Excel file from basic inputs.

        :param title: Title/filename for the Excel file
        :param headers: Comma-separated column headers (e.g., "Product,Jan,Feb,Mar,Total")
        :param data: Pipe-separated rows, comma-separated values (e.g., "Widget,100,120,140|Gadget,150,160,170")
        :param include_totals: Whether to add a totals row with SUM formulas
        :return: HTML with download link
        """
        try:
            header_list = [h.strip() for h in headers.split(",")]

            rows = []
            for row_str in data.split("|"):
                row_values = []
                for val in row_str.split(","):
                    val = val.strip()
                    try:
                        if "." in val:
                            row_values.append(float(val))
                        else:
                            row_values.append(int(val))
                    except ValueError:
                        row_values.append(val)
                rows.append(row_values)

            numeric_cols = []
            if rows:
                for col_idx in range(len(rows[0])):
                    if all(isinstance(row[col_idx], (int, float)) for row in rows if col_idx < len(row)):
                        numeric_cols.append(col_idx)

            spec = {
                "filename": title,
                "sheets": [{
                    "name": title[:31],
                    "headers": header_list,
                    "data": rows,
                    "format": {
                        "bold_headers": True,
                        "auto_fit": True,
                        "freeze_row": 1
                    }
                }]
            }

            if include_totals and numeric_cols:
                totals_row = []
                data_row_count = len(rows)
                start_row = 2
                end_row = start_row + data_row_count - 1

                for col_idx in range(len(header_list)):
                    if col_idx == 0:
                        totals_row.append("TOTAL")
                    elif col_idx in numeric_cols:
                        col_letter = chr(65 + col_idx)
                        totals_row.append(f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
                    else:
                        totals_row.append("")

                spec["sheets"][0]["totals_row"] = totals_row

            return self.create_excel(json.dumps(spec), __user__, __event_emitter__)

        except Exception as e:
            return f"<p style='color:red;'>Error: {str(e)}</p>"
'''

excel_creator_specs = [
    {
        "name": "create_excel",
        "description": "Create an Excel file with multiple sheets, formulas (SUM, AVERAGE, IF, VLOOKUP), and formatting. Use JSON specification for complex reports.",
        "parameters": {
            "type": "object",
            "properties": {
                "specification": {
                    "type": "string",
                    "description": "JSON specification with filename, sheets array containing name, headers, data, totals_row, and format options"
                }
            },
            "required": ["specification"]
        }
    },
    {
        "name": "create_simple_excel",
        "description": "Create a simple Excel file from comma/pipe-separated data. Easier to use for basic spreadsheets.",
        "parameters": {
            "type": "object",
            "properties": {
                "title": {
                    "type": "string",
                    "description": "Title/filename for the Excel file (e.g., 'Q1_Sales')"
                },
                "headers": {
                    "type": "string",
                    "description": "Comma-separated column headers (e.g., 'Product,Jan,Feb,Mar')"
                },
                "data": {
                    "type": "string",
                    "description": "Pipe-separated rows with comma-separated values (e.g., 'Widget,100,120,140|Gadget,150,160,170')"
                },
                "include_totals": {
                    "type": "boolean",
                    "description": "Whether to add a totals row with SUM formulas (default: true)"
                }
            },
            "required": ["title", "headers", "data"]
        }
    }
]

excel_creator_meta = {"description": "Creates Excel files with working formulas, multiple sheets, and professional formatting"}

# ============================================================================
# TOOL 2: Executive Dashboard
# ============================================================================

dashboard_id = "executive_dashboard"
dashboard_name = "Executive Dashboard"

dashboard_content = '''import json
from datetime import datetime
from typing import Optional, Any, List, Dict


class Tools:
    def __init__(self):
        self.plotly_cdn = "https://cdn.plot.ly/plotly-2.27.0.min.js"

    def create_dashboard(
        self,
        specification: str,
        __user__: dict = {},
        __event_emitter__: Any = None
    ) -> str:
        """
        Create an executive dashboard with KPIs and charts.

        :param specification: JSON specification for the dashboard. Format:
            {
                "title": "Q1 2026 Executive Dashboard",
                "subtitle": "Performance Overview",
                "kpis": [
                    {"label": "Revenue", "value": "$1.2M", "change": "+15%", "trend": "up"},
                    {"label": "Users", "value": "45,000", "change": "+8%", "trend": "up"}
                ],
                "charts": [
                    {
                        "type": "line",
                        "title": "Revenue Trend",
                        "labels": ["Jan", "Feb", "Mar"],
                        "datasets": [{"label": "2026", "data": [350000, 420000, 480000]}]
                    },
                    {
                        "type": "pie",
                        "title": "Sales by Region",
                        "labels": ["North", "South", "East", "West"],
                        "data": [35, 25, 22, 18]
                    },
                    {
                        "type": "bar",
                        "title": "Actual vs Target",
                        "labels": ["Q1", "Q2", "Q3", "Q4"],
                        "datasets": [
                            {"label": "Actual", "data": [105, 98, 112, 108]},
                            {"label": "Target", "data": [100, 100, 100, 100]}
                        ]
                    }
                ],
                "theme": "light"
            }
        :return: HTML dashboard
        """
        try:
            if isinstance(specification, str):
                try:
                    spec = json.loads(specification)
                except json.JSONDecodeError as e:
                    return f"<p style='color:red;'>Error: Invalid JSON - {str(e)}</p>"
            else:
                spec = specification

            title = spec.get("title", "Executive Dashboard")
            subtitle = spec.get("subtitle", "")
            kpis = spec.get("kpis", [])
            charts = spec.get("charts", [])
            theme = spec.get("theme", "light")

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

            kpi_html = self._generate_kpi_cards(kpis)
            charts_html = ""
            charts_js = ""
            for i, chart in enumerate(charts):
                chart_id = f"chart_{i}"
                div_html, js_code = self._generate_chart(chart, chart_id)
                charts_html += div_html
                charts_js += js_code

            css = self._get_dashboard_css(theme)

            html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <script src="{self.plotly_cdn}"></script>
    {css}
</head>
<body>
    <div class="dashboard">
        <div class="dashboard-header">
            <div class="dashboard-title">{title}</div>
            {f'<div class="dashboard-subtitle">{subtitle}</div>' if subtitle else ''}
            <div class="dashboard-timestamp">Generated: {timestamp}</div>
        </div>
        {f'<div class="kpi-grid">{kpi_html}</div>' if kpis else ''}
        {f'<div class="chart-grid">{charts_html}</div>' if charts else ''}
        <div class="footer">Generated by Open WebUI Reporting Toolkit</div>
    </div>
    <script>
        document.addEventListener('DOMContentLoaded', function() {{
            {charts_js}
        }});
    </script>
</body>
</html>"""
            return html

        except Exception as e:
            return f"<p style='color:red;'>Error: {str(e)}</p>"

    def _generate_kpi_cards(self, kpis: List[Dict]) -> str:
        cards_html = ""
        for kpi in kpis:
            label = kpi.get("label", "KPI")
            value = kpi.get("value", "0")
            change = kpi.get("change", "")
            trend = kpi.get("trend", "neutral")

            if trend == "up":
                trend_class = "positive"
                trend_arrow = "&#9650;"
            elif trend == "down":
                trend_class = "negative"
                trend_arrow = "&#9660;"
            else:
                trend_class = "neutral"
                trend_arrow = "&#9644;"

            cards_html += f"""
            <div class="kpi-card {trend_class}">
                <div class="kpi-label">{label}</div>
                <div class="kpi-value">{value}</div>
                {f'<div class="kpi-change"><span class="trend-arrow">{trend_arrow}</span> {change}</div>' if change else ''}
            </div>"""
        return cards_html

    def _generate_chart(self, chart: Dict, chart_id: str) -> tuple:
        chart_type = chart.get("type", "bar")
        title = chart.get("title", "Chart")

        div_html = f"""
        <div class="chart-container">
            <div class="chart-title">{title}</div>
            <div id="{chart_id}" class="chart"></div>
        </div>"""

        if chart_type == "bar":
            js_code = self._generate_bar_chart_js(chart, chart_id)
        elif chart_type == "line":
            js_code = self._generate_line_chart_js(chart, chart_id)
        elif chart_type == "pie":
            js_code = self._generate_pie_chart_js(chart, chart_id)
        else:
            js_code = self._generate_bar_chart_js(chart, chart_id)

        return div_html, js_code

    def _generate_bar_chart_js(self, chart: Dict, chart_id: str) -> str:
        labels = chart.get("labels", [])
        datasets = chart.get("datasets", [])
        traces = []
        colors = ['#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899']

        for i, dataset in enumerate(datasets):
            trace = {
                "x": labels,
                "y": dataset.get("data", []),
                "name": dataset.get("label", f"Series {i+1}"),
                "type": "bar",
                "marker": {"color": colors[i % len(colors)]}
            }
            traces.append(trace)

        layout = {"barmode": "group", "margin": {"t": 20, "r": 20, "b": 40, "l": 50}, "legend": {"orientation": "h", "y": -0.15}, "paper_bgcolor": "rgba(0,0,0,0)", "plot_bgcolor": "rgba(0,0,0,0)"}
        config = {"responsive": True, "displayModeBar": False}
        return f'Plotly.newPlot("{chart_id}", {json.dumps(traces)}, {json.dumps(layout)}, {json.dumps(config)});'

    def _generate_line_chart_js(self, chart: Dict, chart_id: str) -> str:
        labels = chart.get("labels", [])
        datasets = chart.get("datasets", [])
        traces = []
        colors = ['#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899']

        for i, dataset in enumerate(datasets):
            trace = {
                "x": labels,
                "y": dataset.get("data", []),
                "name": dataset.get("label", f"Series {i+1}"),
                "type": "scatter",
                "mode": "lines+markers",
                "line": {"color": colors[i % len(colors)], "width": 3},
                "marker": {"size": 8}
            }
            traces.append(trace)

        layout = {"margin": {"t": 20, "r": 20, "b": 40, "l": 50}, "legend": {"orientation": "h", "y": -0.15}, "paper_bgcolor": "rgba(0,0,0,0)", "plot_bgcolor": "rgba(0,0,0,0)", "xaxis": {"showgrid": True, "gridcolor": "#e5e7eb"}, "yaxis": {"showgrid": True, "gridcolor": "#e5e7eb"}}
        config = {"responsive": True, "displayModeBar": False}
        return f'Plotly.newPlot("{chart_id}", {json.dumps(traces)}, {json.dumps(layout)}, {json.dumps(config)});'

    def _generate_pie_chart_js(self, chart: Dict, chart_id: str) -> str:
        labels = chart.get("labels", [])
        data = chart.get("data", [])
        colors = ['#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899', '#14b8a6', '#f97316']

        trace = {
            "labels": labels,
            "values": data,
            "type": "pie",
            "hole": 0.4,
            "marker": {"colors": colors[:len(labels)]},
            "textinfo": "label+percent",
            "textposition": "outside"
        }

        layout = {"margin": {"t": 20, "r": 20, "b": 20, "l": 20}, "showlegend": True, "legend": {"orientation": "h", "y": -0.1}, "paper_bgcolor": "rgba(0,0,0,0)"}
        config = {"responsive": True, "displayModeBar": False}
        return f'Plotly.newPlot("{chart_id}", [{json.dumps(trace)}], {json.dumps(layout)}, {json.dumps(config)});'

    def _get_dashboard_css(self, theme: str = "light") -> str:
        if theme == "dark":
            bg_color = "#1f2937"
            text_color = "#f9fafb"
            card_bg = "#374151"
        else:
            bg_color = "#f3f4f6"
            text_color = "#111827"
            card_bg = "#ffffff"

        return f"""<style>
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:{bg_color};color:{text_color};}}
.dashboard{{max-width:1400px;margin:0 auto;padding:30px;}}
.dashboard-header{{text-align:center;margin-bottom:40px;}}
.dashboard-title{{font-size:32px;font-weight:700;margin-bottom:8px;}}
.dashboard-subtitle{{font-size:18px;color:#6b7280;margin-bottom:8px;}}
.dashboard-timestamp{{font-size:14px;color:#9ca3af;}}
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:24px;margin-bottom:40px;}}
.kpi-card{{background:{card_bg};border-radius:16px;padding:24px;box-shadow:0 4px 6px rgba(0,0,0,0.1);text-align:center;}}
.kpi-label{{font-size:14px;font-weight:500;color:#6b7280;text-transform:uppercase;margin-bottom:8px;}}
.kpi-value{{font-size:36px;font-weight:700;margin-bottom:8px;}}
.kpi-change{{font-size:14px;font-weight:600;display:flex;align-items:center;justify-content:center;gap:4px;}}
.kpi-card.positive .kpi-change{{color:#10b981;}}
.kpi-card.negative .kpi-change{{color:#ef4444;}}
.kpi-card.neutral .kpi-change{{color:#6b7280;}}
.chart-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(400px,1fr));gap:24px;margin-bottom:40px;}}
.chart-container{{background:{card_bg};border-radius:16px;padding:24px;box-shadow:0 4px 6px rgba(0,0,0,0.1);}}
.chart-title{{font-size:18px;font-weight:600;margin-bottom:16px;text-align:center;}}
.chart{{min-height:300px;width:100%;}}
.footer{{text-align:center;padding:20px;color:#9ca3af;font-size:12px;}}
</style>"""

    def create_simple_dashboard(
        self,
        title: str,
        kpis: str,
        chart_type: str = "bar",
        chart_title: str = "Chart",
        chart_labels: str = "",
        chart_data: str = "",
        __user__: dict = {},
        __event_emitter__: Any = None
    ) -> str:
        """
        Create a simple dashboard from basic inputs.

        :param title: Dashboard title
        :param kpis: KPIs in format "Label:Value:Change:trend|..." (e.g., "Revenue:$1.2M:+15%:up|Users:45K:+8%:up")
        :param chart_type: Type of chart (bar, line, pie)
        :param chart_title: Title for the chart
        :param chart_labels: Comma-separated labels (e.g., "Jan,Feb,Mar,Apr")
        :param chart_data: Comma-separated data values (e.g., "100,120,140,160")
        :return: HTML dashboard
        """
        try:
            kpi_list = []
            if kpis:
                for kpi_str in kpis.split("|"):
                    parts = kpi_str.split(":")
                    if len(parts) >= 2:
                        kpi_list.append({
                            "label": parts[0].strip(),
                            "value": parts[1].strip(),
                            "change": parts[2].strip() if len(parts) > 2 else "",
                            "trend": parts[3].strip() if len(parts) > 3 else "neutral"
                        })

            labels = [l.strip() for l in chart_labels.split(",")] if chart_labels else []
            data = []
            if chart_data:
                for val in chart_data.split(","):
                    try:
                        data.append(float(val.strip()))
                    except ValueError:
                        data.append(0)

            spec = {"title": title, "kpis": kpi_list, "charts": []}

            if labels and data:
                if chart_type == "pie":
                    spec["charts"].append({"type": "pie", "title": chart_title, "labels": labels, "data": data})
                else:
                    spec["charts"].append({"type": chart_type, "title": chart_title, "labels": labels, "datasets": [{"label": "Value", "data": data}]})

            return self.create_dashboard(json.dumps(spec), __user__, __event_emitter__)

        except Exception as e:
            return f"<p style='color:red;'>Error: {str(e)}</p>"
'''

dashboard_specs = [
    {
        "name": "create_dashboard",
        "description": "Create an executive dashboard with KPI cards and interactive charts (bar, line, pie). Use JSON specification for complex dashboards.",
        "parameters": {
            "type": "object",
            "properties": {
                "specification": {
                    "type": "string",
                    "description": "JSON specification with title, subtitle, kpis array, charts array, and theme"
                }
            },
            "required": ["specification"]
        }
    },
    {
        "name": "create_simple_dashboard",
        "description": "Create a simple dashboard from basic pipe/comma-separated inputs. Easier to use for quick dashboards.",
        "parameters": {
            "type": "object",
            "properties": {
                "title": {
                    "type": "string",
                    "description": "Dashboard title"
                },
                "kpis": {
                    "type": "string",
                    "description": "KPIs in format 'Label:Value:Change:trend|...' (e.g., 'Revenue:$1.2M:+15%:up|Users:45K:+8%:up')"
                },
                "chart_type": {
                    "type": "string",
                    "description": "Type of chart: bar, line, or pie"
                },
                "chart_title": {
                    "type": "string",
                    "description": "Title for the chart"
                },
                "chart_labels": {
                    "type": "string",
                    "description": "Comma-separated labels (e.g., 'Jan,Feb,Mar,Apr')"
                },
                "chart_data": {
                    "type": "string",
                    "description": "Comma-separated data values (e.g., '100,120,140,160')"
                }
            },
            "required": ["title", "kpis"]
        }
    }
]

dashboard_meta = {"description": "Creates executive dashboards with KPI cards and interactive Plotly.js charts"}

# ============================================================================
# DATABASE INSERTION
# ============================================================================

print("Connecting to PostgreSQL...")

try:
    import psycopg2
except ImportError:
    import subprocess
    subprocess.run(["pip", "install", "psycopg2-binary", "-q"])
    import psycopg2

conn = psycopg2.connect(DB_URL)
cursor = conn.cursor()

# Clean up old entries
cursor.execute('DELETE FROM function WHERE id IN (%s, %s)', (excel_creator_id, dashboard_id))
cursor.execute('DELETE FROM tool WHERE id IN (%s, %s)', (excel_creator_id, dashboard_id))
print("Cleaned up old entries")

# Insert Excel Creator
print(f"Inserting tool: {excel_creator_name}...")
cursor.execute('''
    INSERT INTO tool (id, user_id, name, content, specs, meta, created_at, updated_at, valves, access_control)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
''', (
    excel_creator_id,
    user_id,
    excel_creator_name,
    excel_creator_content,
    json.dumps(excel_creator_specs),
    json.dumps(excel_creator_meta),
    now,
    now,
    '{}',
    json.dumps({"read": {"group_ids": [], "user_ids": []}, "write": {"group_ids": [], "user_ids": []}})
))

# Insert Executive Dashboard
print(f"Inserting tool: {dashboard_name}...")
cursor.execute('''
    INSERT INTO tool (id, user_id, name, content, specs, meta, created_at, updated_at, valves, access_control)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
''', (
    dashboard_id,
    user_id,
    dashboard_name,
    dashboard_content,
    json.dumps(dashboard_specs),
    json.dumps(dashboard_meta),
    now,
    now,
    '{}',
    json.dumps({"read": {"group_ids": [], "user_ids": []}, "write": {"group_ids": [], "user_ids": []}})
))

conn.commit()
print("Committed!")

# Verify
cursor.execute('SELECT id, name FROM tool')
rows = cursor.fetchall()
print(f"Tools in database: {rows}")

conn.close()
print("Done! Reporting Toolkit installed successfully.")
