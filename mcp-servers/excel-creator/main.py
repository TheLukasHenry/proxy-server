"""
Excel Creator MCP Server
Creates Excel files with data, formulas, and charts.
Exposes tools via OpenAPI for the MCP proxy.
"""
from fastapi import FastAPI, HTTPException
from fastapi.responses import JSONResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Any
import base64
import io

app = FastAPI(
    title="Excel Creator MCP",
    description="Create Excel spreadsheets with data, formulas, and charts",
    version="1.0.0"
)


class SimpleExcelRequest(BaseModel):
    """Request for creating a simple Excel file."""
    title: str = Field(..., description="Name for the Excel file (without .xlsx)")
    headers: str = Field(..., description="Comma-separated column headers (e.g., 'Product,Jan,Feb,Mar')")
    data: str = Field(..., description="Pipe-separated rows with comma-separated values (e.g., 'Widget,100,150,200|Gadget,250,300,350')")
    include_totals: bool = Field(default=True, description="Add a TOTAL row with SUM formulas")


class AdvancedExcelRequest(BaseModel):
    """Request for creating an advanced Excel file with multiple sheets."""
    title: str = Field(..., description="Name for the Excel file")
    sheets: List[dict] = Field(..., description="List of sheet configurations with name, headers, and data")
    include_charts: bool = Field(default=False, description="Add charts to visualize the data")


class ExcelResponse(BaseModel):
    """Response containing the Excel file."""
    success: bool
    filename: str
    download_html: str
    message: str
    rows: int
    columns: int


@app.get("/health")
async def health():
    """Health check endpoint."""
    return {"status": "healthy", "service": "excel-creator"}


@app.post("/create_simple_excel", response_model=ExcelResponse)
async def create_simple_excel(request: SimpleExcelRequest):
    """
    Create a simple Excel file from comma/pipe-separated data.

    Example:
        title: "Q1Sales"
        headers: "Product,Jan,Feb,Mar"
        data: "Widget,100,150,200|Gadget,250,300,350"

    Returns an HTML download link for the generated Excel file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    # Parse headers
    header_list = [h.strip() for h in request.headers.split(",")]

    # Parse data rows
    rows = []
    for row_str in request.data.split("|"):
        row = []
        for val in row_str.split(","):
            val = val.strip()
            try:
                val = int(val)
            except ValueError:
                try:
                    val = float(val)
                except ValueError:
                    pass
            row.append(val)
        rows.append(row)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = str(request.title)[:31]

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(header_list, 1):
        cell = ws.cell(row=1, column=col, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border

    # Add totals row if requested
    if request.include_totals and rows:
        total_row = len(rows) + 2
        total_cell = ws.cell(row=total_row, column=1, value="TOTAL")
        total_cell.font = Font(bold=True)
        total_cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        total_cell.border = border

        for col in range(2, len(header_list) + 1):
            col_letter = get_column_letter(col)
            formula = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
            cell = ws.cell(row=total_row, column=col, value=formula)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            cell.border = border

    # Auto-adjust column widths
    for col in range(1, len(header_list) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Encode as base64
    b64 = base64.b64encode(buffer.read()).decode()
    fname = f"{request.title}.xlsx"

    # Create HTML download link
    download_html = f'''<div style="padding:20px;background:linear-gradient(135deg,#1e3a5f,#2d5a87);border-radius:12px;color:white;margin:10px 0;">
<h3 style="margin:0 0 10px 0;">Excel File Created: {fname}</h3>
<p style="margin:0 0 15px 0;opacity:0.9;">Contains {len(rows)} data rows with {len(header_list)} columns{" and SUM formulas" if request.include_totals else ""}</p>
<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{fname}" style="display:inline-block;padding:12px 24px;background:#4CAF50;color:white;text-decoration:none;border-radius:6px;font-weight:bold;">Download {fname}</a>
</div>'''

    return ExcelResponse(
        success=True,
        filename=fname,
        download_html=download_html,
        message=f"Created {fname} with {len(rows)} rows",
        rows=len(rows),
        columns=len(header_list)
    )


@app.post("/create_excel_from_json")
async def create_excel_from_json(
    title: str = "Report",
    headers: List[str] = None,
    data: List[List[Any]] = None,
    include_totals: bool = True
):
    """
    Create an Excel file from JSON arrays (alternative format).

    Example:
        title: "Q1Sales"
        headers: ["Product", "Jan", "Feb", "Mar"]
        data: [["Widget", 100, 150, 200], ["Gadget", 250, 300, 350]]
    """
    if not headers or not data:
        raise HTTPException(status_code=400, detail="headers and data are required")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = str(title)[:31]

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border

    # Add totals
    if include_totals and data:
        total_row = len(data) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        for col in range(2, len(headers) + 1):
            col_letter = get_column_letter(col)
            formula = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
            ws.cell(row=total_row, column=col, value=formula).font = Font(bold=True)

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save and encode
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    b64 = base64.b64encode(buffer.read()).decode()
    fname = f"{title}.xlsx"

    download_html = f'''<div style="padding:20px;background:linear-gradient(135deg,#1e3a5f,#2d5a87);border-radius:12px;color:white;margin:10px 0;">
<h3 style="margin:0 0 10px 0;">Excel File Created: {fname}</h3>
<p style="margin:0 0 15px 0;opacity:0.9;">Contains {len(data)} data rows with {len(headers)} columns</p>
<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{fname}" style="display:inline-block;padding:12px 24px;background:#4CAF50;color:white;text-decoration:none;border-radius:6px;font-weight:bold;">Download {fname}</a>
</div>'''

    return {
        "success": True,
        "filename": fname,
        "download_html": download_html,
        "message": f"Created {fname} with {len(data)} rows",
        "rows": len(data),
        "columns": len(headers)
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
